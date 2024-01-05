import os
import sys
import win32com.client
from student import *

from tkinter import messagebox
import time

import openpyxl
# from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font


def print_debug(debug, message):
    if debug:
        print(message)

def listFiles(directory, extension=".docx"):
    """
    :param directory: directory to parse
    :param extension: extension to look for
    :return: listFiles, array with files
    """
    # list all files
    files = os.listdir(directory)
    listFiles = []
    for file in files:
        if file.endswith(extension):
            listFiles.append(file)
    return listFiles
    # return (files, listFiles)


def ensure_file_is_closed_and_exists(file, debug=False):
    command_executed = False
    error = False
    while not command_executed:
        if os.path.exists(file):
            try:
                os.rename(file, file)
                print_debug(debug, 'Access on file "' + file + '" is available!')
                time.sleep(1)
                if error:
                    ppt_app = win32com.client.Dispatch("PowerPoint.Application")
                command_executed = True
            except OSError as e:
                print('Access-error on file "' + file + '"! \n' + str(e))
                messagebox.showinfo(title="Script de correction automatique",
                                    message="Fermer la presentation " + file + " pour la nouvelle correction")
                error = True
        else:
            print("file don't exist", file)
            exit(2)


def open_presentation(ppt_app, file, debug=False):

    # self.app = win32com.client.Dispatch("PowerPoint.Application")
    # Open presentation
    presentation = ""
    print_debug(debug, file)
    filename = os.path.abspath(file)
    print_debug(debug, filename)
    try:
        presentation = ppt_app.Presentations.Open(filename)
    except Exception as e:
        print("erreur dans l'ouverture de la presentation" + str(e))
        quit(2)

    print_debug(debug, "ok, presentation open")
    return presentation


def close_powerpoint(debug):
    try:
        # Create a Word application object
        ppt_app = win32com.client.Dispatch("Powerpoint.Application")
        # quit without saving
        ppt_app.Quit(SaveChanges=0)

        print_debug(debug, "Ppt closed successfully")

    except Exception as e:
        print(f"Error: {e}")

def fill_first_lines_excel(worksheet, student):
    row = 1
    worksheet.cell(row=row, column=1).value = "Nom"
    worksheet.cell(row=row, column=1).font = Font(bold=True)
    worksheet.cell(row=row, column=2).value = "Prénom"
    worksheet.cell(row=row, column=2).font = Font(bold=True)
    worksheet.cell(row=row, column=3).value = "Total"
    worksheet.cell(row=row, column=3).font = Font(bold=True)
    col = 4
    #print(str(student))
    #for key in student.scores.keys():
    for key in student.max_points.keys():
        worksheet.cell(row=row, column=col).value = key
        worksheet.cell(row=row, column=col).font = Font(bold=True)
        col += 1
    worksheet.cell(row=row, column=col).value = "à vérifier manuellement"
    worksheet.cell(row=row, column=col).font = Font(bold=True)
    col = 4
    row += 1
    for key, value in student.max_points.items():
        worksheet.cell(row=row, column=col).value = key
        worksheet.cell(row=row, column=col).font = Font(bold=True)
        worksheet.cell(row=row, column=col).font = Font(italic=True)
        col += 1
    row += 1
    col = 4
    for key, value in student.max_points.items():
        worksheet.cell(row=row, column=col).value = value
        worksheet.cell(row=row, column=col).font = Font(bold=True)
        col += 1
    worksheet.cell(row=row, column=3).value = "=sum(" + \
                                              get_column_letter(4) + str(row) + \
                                              ":" + get_column_letter(4 + len(student.max_points.items())) + \
                                              str(row) + ")"

    worksheet.cell(row=row, column=col).font = Font(bold=True)

    row += 1

    worksheet.freeze_panes = 'D4'
    # col += 2
    # for key in reasons_set:
    #     worksheet.cell(row=1, column=col).value = key
    #     col += 1
    return row


def fill_result_line_in_excel(worksheet, row, student):
    # Pour chaque élément du set, l'ajouter dans une nouvelle cellule
    worksheet.cell(row=row, column=1).value = student.name.capitalize()
    worksheet.cell(row=row, column=2).value = student.firstname.capitalize()
    worksheet.cell(row=row, column=3).value \
        = "=sum(" + get_column_letter(4) + str(row) + ":" + get_column_letter(4 + len(student.scores.items())) + str(
        row) + ")"
    col = 4
    # print("à vérifier : ", student.to_check)
    # print("sytles : ", student.scores["styles"], "et liens : ", student.scores["lien"])
    # TODO add conditional formatting : https://openpyxl.readthedocs.io/en/latest/formatting.html
    #        --> < max_score/2 --> red font color
    #        --> = max_score   --> green font color
    # Todo add formulas : https://openpyxl.readthedocs.io/en/latest/usage.html?highlight=formula#using-formulae
    for key, value in student.scores.items():
        worksheet.cell(row=row, column=col).value = value
        cell = get_column_letter(col) + str(row)
        blank_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        worksheet[cell].fill = blank_fill
#was        blank_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # Create the comment with the "Why that score" if any
        why = student.reasons[key]
        if why != "":
            comment = Comment(why, "François Schoubben")
            worksheet[cell].comment = comment
        if key in student.to_check:
            # print("mettre", key, " en jaune")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            worksheet[cell].fill = yellow_fill
        col += 1

    worksheet.cell(row=row, column=col).value = student.to_check_manually


def fill_last_line_in_excel(worksheet, row, student, number_of_non_student_lines):
    row += 1
    worksheet.cell(row=row, column=1).value = "Moyenne étudiant"
    for col in range(3, len(student.max_points) + 3):
        worksheet.cell(row=row, column=col).value = \
            "=average(" + get_column_letter(col) + str(number_of_non_student_lines + 1) \
            + ":" + get_column_letter(col) + str(row - 1) + ")"

    row += 1
    worksheet.cell(row=row, column=1).value = "Min étudiant"
    for col in range(3, len(student.max_points) + 3):
        worksheet.cell(row=row, column=col).value = \
            "=min(" + get_column_letter(col) + str(number_of_non_student_lines + 1) \
            + ":" + get_column_letter(col) + str(row - 2) + ")"
    row += 1
    worksheet.cell(row=row, column=1).value = "MAX étudiant"
    for col in range(3, len(student.max_points) + 3):
        worksheet.cell(row=row, column=col).value = \
            "=max(" + get_column_letter(col) + str(number_of_non_student_lines + 1) \
            + ":" + get_column_letter(col) + str(row - 3) + ")"
def create_xls_sheets(student, groups):
    # Créer un nouveau tableur
    workbook = openpyxl.Workbook()
    # Créer une nouvelle feuille par groupe
    worksheets={}
    first_empty_row={}
    # rows={}
    for group in groups:
        worksheet = workbook.create_sheet(group)
        worksheets[group]=worksheet
        first_empty_row[group] = fill_first_lines_excel(worksheet, student)
       # row[group] = first_empty_row[group]
    #if "Unknown" not in groups:
    #        worksheets.append(workbook.create_sheet("Unknown"))
    workbook.remove(workbook["Sheet"])
    return workbook, worksheets, first_empty_row

def execute_ensuring_file_not_open(file, command, debug=False):
    command_executed = False
    while not command_executed:
        if os.path.exists(file):
            try:
                os.rename(file, file)
                print_debug(debug, 'Access on file "' + file + '" is available!')
                command(file)
                command_executed = True
            except OSError as e:
                print('Access-error on file "' + file + '"! \n' + str(e))
                messagebox.showinfo(title="Script de correction automatique",
                                    message="Fermer le document Excel pour la nouvelle correction")
        else:
            command(file)

def save_in_excel_file(excel_file_for_results, students, groups):
    (workbook,worksheets, first_empty_rows) = create_xls_sheets(Student, groups)
    rows = dict(first_empty_rows)
    for student in students:
        fill_result_line_in_excel(worksheets[student.group], rows[student.group], student)
        rows[student.group] += 1
    for key in groups:
        fill_last_line_in_excel(worksheets[key], rows[key], student, first_empty_rows[key] - 1)

    # Enregistrer le tableur
    execute_ensuring_file_not_open(excel_file_for_results, workbook.save)
def verifier_nom_fichiers(mfile, template, student):
    # TODO NextVersion (manually done 2023): vérifier espaces ! (-1)
    raisons = ""
    points = 0
    if mfile.startswith(template):
        points += student.max_points["nomFichiers"]
    #mfile = mfile.replace("—", "")
    #mfile = mfile.replace(" ", "")
    # if "schoubben" in mfile:  # TODO 2023 : manually done, try to emprove to automate...
    #     raisons="mauvais nom de fichier; "
    nomcomplet = mfile.split("-")
    nom = nomcomplet[3]
    prenom = nomcomplet[4]
    # print("in nom prenom = ", nom, prenom, nomcomplet)
    student.name = nom
    student.firstname = prenom
    student.scores["nomFichiers"] = points
    student.reasons["nomFichiers"] = raisons
    return student.max_points["nomFichiers"]


# 2 formats : word/opendoc + pdf /2 TODO : améliorer pour vérifier types
def verifier_deux_formats_fichiers(filename, liste_fichiers, max_points, scores_set, reasons_set, key="format"):
    # f : filename (string)
    nb_fichiers = 0
    for el in liste_fichiers:
        if filename[0:-4] in el:
            nb_fichiers += 1
    # print("nb fichiers", nbFichiers)
    if nb_fichiers == 2:
        scores_set[key] = max_points
        reasons_set[key] = ""
    else:
        scores_set[key] = 0
        reasons_set[key] = "il n'y a pas les 2 formats de fichier"


# moins de 3Mo (moins de 1Mo) /2
def verifier_moins_de_3_mo(filename, max_size, max_points):
    # f : filename (string)
    try:
        file_info = os.stat(filename)
        if file_info.st_size < max_size * 1000000:
            return max_points, ""
        else:
            return 0, "fichier trop gros"
    except Exception as e:
        sys.stderr.write("erreur de nom de fichier" + str(e))



def main():
    ppt_app = win32com.client.Dispatch("PowerPoint.Application")
    files = listFiles(".", ".pptx")
    print(files)
    presentation = open_presentation(ppt_app, files[0], debug=True)

    ppt_app.Quit()
if __name__ == "__main__":
    main()
